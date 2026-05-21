import os
import re
from datetime import datetime, date, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.database import SessionLocal
from app.models.playlist import Playlist
from app.models.follower_history import FollowerHistory

# Usage:
#   cd "C:\\Nerd Engine\\spotify-growth-hub"
#   $env:DRY_RUN="true"
#   python import_history_from_excel.py
#
# Optional custom Excel path:
#   $env:EXCEL_FILE="C:\\path\\to\\Spotify_Master.xlsx"
#   python import_history_from_excel.py

EXCEL_FILE = Path(os.getenv("EXCEL_FILE", "Spotify_Master.xlsx"))
DRY_RUN = os.getenv("DRY_RUN", "false").lower() in {"1", "true", "yes", "y"}

# Keep existing DB data for these two dates. Excel values for these dates are skipped.
PROTECTED_DATES = {date(2026, 5, 16), date(2026, 5, 17)}

# If your Excel date headers are like 13/5 with no year, use this year.
DEFAULT_IMPORT_YEAR = int(os.getenv("DEFAULT_IMPORT_YEAR", "2026"))

BATCH_SIZE = 1000


def norm(value: Any) -> str:
    return str(value or "").strip()


def header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", norm(value).lower()).strip("_")


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = norm(value)
    if not text:
        return None

    text = text.replace("-", "/").replace(".", "/")

    # YYYY/MM/DD
    match = re.fullmatch(r"(20\d{2})/(\d{1,2})/(\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        try:
            return date(year, month, day)
        except ValueError:
            return None

    # DD/MM/YYYY or DD/MM/YY. Your sheets use day/month dates such as 16/5/2026.
    match = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})", text)
    if match:
        day, month, year = match.groups()
        year_int = int(year)
        if year_int < 100:
            year_int += 2000
        try:
            return date(year_int, int(month), int(day))
        except ValueError:
            return None

    # DD/MM with no year.
    match = re.fullmatch(r"(\d{1,2})/(\d{1,2})", text)
    if match:
        day, month = map(int, match.groups())
        try:
            return date(DEFAULT_IMPORT_YEAR, month, day)
        except ValueError:
            return None

    return None


def extract_spotify_id(value: Any) -> str | None:
    text = norm(value)
    if not text:
        return None

    patterns = [
        r"open\.spotify\.com/playlist/([A-Za-z0-9]+)",
        r"spotify:playlist:([A-Za-z0-9]+)",
        r"playlist/([A-Za-z0-9]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)

    # Spotify playlist IDs are usually 22 chars. Allow 16+ to support exports.
    if re.fullmatch(r"[A-Za-z0-9]{16,}", text):
        return text

    return None


def parse_int(value: Any) -> int | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = norm(value)
    if not text or text in {"-", "—", "n/a", "N/A"}:
        return None

    text = text.replace(",", "").replace("+", "")
    try:
        return int(float(text))
    except Exception:
        return None


def row_history_date(row: FollowerHistory) -> date | None:
    value = getattr(row, "date", None)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    created_at = getattr(row, "created_at", None)
    if isinstance(created_at, datetime):
        return created_at.date()
    if isinstance(created_at, date):
        return created_at

    return None


def build_playlist_indexes(db):
    playlists = db.query(Playlist).all()

    by_db_id: dict[int, int] = {}
    by_spotify: dict[str, int] = {}
    by_name: dict[str, int] = {}

    for playlist in playlists:
        playlist_id = int(playlist.id)
        by_db_id[playlist_id] = playlist_id

        name_key = norm(getattr(playlist, "name", None)).lower()
        if name_key and name_key not in by_name:
            by_name[name_key] = playlist_id

        for value in [
            getattr(playlist, "spotify_id", None),
            getattr(playlist, "spotify_playlist_id", None),
            getattr(playlist, "spotify_url", None),
            getattr(playlist, "playlist_url", None),
            getattr(playlist, "url", None),
            getattr(playlist, "external_url", None),
        ]:
            spotify_id = extract_spotify_id(value)
            if spotify_id and spotify_id not in by_spotify:
                by_spotify[spotify_id] = playlist_id

    return by_db_id, by_spotify, by_name


def resolve_playlist_id(row_data: dict[str, Any], by_db_id, by_spotify, by_name) -> int | None:
    # First try true database playlist IDs.
    db_id_keys = [
        "db_playlist_id",
        "database_playlist_id",
        "internal_playlist_id",
        "history_playlist_id",
        "database_id",
    ]
    for key in db_id_keys:
        parsed_id = parse_int(row_data.get(key))
        if parsed_id and parsed_id in by_db_id:
            return by_db_id[parsed_id]

    # Many exports call the Spotify ID column "PlaylistID" / "playlistid" / "playlist_id". So:
    # 1) if numeric and exists in DB, use it as DB ID
    # 2) otherwise treat it as Spotify playlist ID/link
    flexible_id_keys = ["playlistid", "playlist_id", "playlist", "id"]
    for key in flexible_id_keys:
        value = row_data.get(key)

        parsed_id = parse_int(value)
        if parsed_id and parsed_id in by_db_id:
            return by_db_id[parsed_id]

        spotify_id = extract_spotify_id(value)
        if spotify_id and spotify_id in by_spotify:
            return by_spotify[spotify_id]

    # Other possible Spotify ID/link columns.
    spotify_id_keys = [
        "spotify_id",
        "spotify_playlist_id",
        "playlist_spotify_id",
        "spotify_playlist",
        "spotify_link",
        "playlist_link",
        "playlist_url",
        "spotify_url",
        "url",
        "link",
        "uri",
    ]
    for key in spotify_id_keys:
        spotify_id = extract_spotify_id(row_data.get(key))
        if spotify_id and spotify_id in by_spotify:
            return by_spotify[spotify_id]

    # Last fallback: playlist name.
    name = (
        row_data.get("title")
        or row_data.get("playlist")
        or row_data.get("playlist_name")
        or row_data.get("name")
    )
    name_key = norm(name).lower()
    if name_key:
        return by_name.get(name_key)

    return None


def find_header_row(sheet):
    """Find the first row that looks like headers: it must contain at least one date column."""
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        date_count = sum(1 for value in row if parse_date(value))
        non_empty_count = sum(1 for value in row if norm(value))
        if date_count >= 1 and non_empty_count >= 2:
            return row_index, row
    return None, None


def collect_excel_history_rows(excel_file: Path, db):
    by_db_id, by_spotify, by_name = build_playlist_indexes(db)
    print(f"Loaded DB playlist indexes: {len(by_db_id)} DB IDs, {len(by_spotify)} Spotify IDs, {len(by_name)} names")
    wb = load_workbook(excel_file, data_only=True, read_only=True)

    imported_by_key: dict[tuple[int, date], int] = {}
    skipped_unmatched_rows = 0
    skipped_protected_cells = 0
    skipped_invalid_values = 0
    sheets_with_dates = 0
    unmatched_examples = []

    for sheet in wb.worksheets:
        header_row_index, headers_raw = find_header_row(sheet)
        if not headers_raw:
            continue

        headers = [header_key(value) for value in headers_raw]
        date_cols = {
            index: parsed_day
            for index, raw_value in enumerate(headers_raw)
            if (parsed_day := parse_date(raw_value)) is not None
        }

        if not date_cols:
            continue

        sheets_with_dates += 1
        print(f"Sheet '{sheet.title}': header row {header_row_index}, date columns: " + ", ".join(day.isoformat() for day in date_cols.values()))
        print("  Headers:", ", ".join([h for h in headers if h][:20]))

        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not any(norm(value) for value in row):
                continue

            row_data = {
                headers[i]: row[i]
                for i in range(min(len(headers), len(row)))
                if headers[i]
            }

            playlist_id = resolve_playlist_id(row_data, by_db_id, by_spotify, by_name)
            if not playlist_id:
                skipped_unmatched_rows += 1
                if len(unmatched_examples) < 8:
                    unmatched_examples.append({
                        "sheet": sheet.title,
                        "playlistid": row_data.get("playlistid"),
                        "playlist_id": row_data.get("playlist_id"),
                        "id": row_data.get("id"),
                        "spotify_id": row_data.get("spotify_id"),
                        "name": row_data.get("name") or row_data.get("playlist") or row_data.get("playlist_name") or row_data.get("title"),
                    })
                continue

            for col_index, day in date_cols.items():
                if col_index >= len(row):
                    continue

                followers = parse_int(row[col_index])
                if followers is None:
                    skipped_invalid_values += 1
                    continue

                if day in PROTECTED_DATES:
                    skipped_protected_cells += 1
                    continue

                # If duplicate cells exist across sheets, the last one wins.
                imported_by_key[(playlist_id, day)] = followers

    return {
        "imported_by_key": imported_by_key,
        "skipped_unmatched_rows": skipped_unmatched_rows,
        "skipped_protected_cells": skipped_protected_cells,
        "skipped_invalid_values": skipped_invalid_values,
        "sheets_with_dates": sheets_with_dates,
        "unmatched_examples": unmatched_examples,
    }


def delete_replace_history(db, imported_by_key: dict[tuple[int, date], int]):
    playlist_ids = sorted({playlist_id for playlist_id, _day in imported_by_key.keys()})
    if not playlist_ids:
        return 0, 0

    existing_rows = db.query(FollowerHistory).filter(FollowerHistory.playlist_id.in_(playlist_ids)).all()

    deleted_count = 0
    for row in existing_rows:
        day = row_history_date(row)
        if day in PROTECTED_DATES:
            continue
        db.delete(row)
        deleted_count += 1

    batch = []
    inserted_count = 0

    for (playlist_id, day), followers in sorted(imported_by_key.items(), key=lambda item: (item[0][0], item[0][1])):
        batch.append(
            FollowerHistory(
                playlist_id=playlist_id,
                date=day,
                followers=followers,
                created_at=datetime.combine(day, time.min),
            )
        )
        inserted_count += 1

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            batch.clear()

    if batch:
        db.bulk_save_objects(batch)

    return deleted_count, inserted_count


def main():
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Missing Excel file: {EXCEL_FILE.resolve()}")

    print(f"Excel file: {EXCEL_FILE.resolve()}")
    print("Mode:", "DRY RUN - no database changes" if DRY_RUN else "LIVE - database will be updated")
    print("Protected dates:", ", ".join(sorted(day.isoformat() for day in PROTECTED_DATES)))
    print(f"Default year for date headers without year: {DEFAULT_IMPORT_YEAR}")

    db = SessionLocal()
    try:
        result = collect_excel_history_rows(EXCEL_FILE, db)
        imported_by_key = result["imported_by_key"]

        print(f"Sheets with history date columns: {result['sheets_with_dates']}")
        print(f"Matched Excel history cells ready to import: {len(imported_by_key)}")
        print(f"Skipped unmatched Excel rows: {result['skipped_unmatched_rows']}")
        print(f"Skipped protected-date Excel cells: {result['skipped_protected_cells']}")
        print(f"Skipped invalid/empty history cells: {result['skipped_invalid_values']}")

        if result["unmatched_examples"]:
            print("Sample unmatched rows:")
            for item in result["unmatched_examples"]:
                print(" ", item)

        deleted_count, inserted_count = delete_replace_history(db, imported_by_key)
        print(f"Database rows to delete, excluding protected dates: {deleted_count}")
        print(f"Database rows to insert from Excel: {inserted_count}")

        if DRY_RUN:
            db.rollback()
            print("Dry run complete. No changes were saved.")
        else:
            db.commit()
            print("Import complete. Existing history was replaced, while 2026-05-16 and 2026-05-17 stayed untouched.")

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
